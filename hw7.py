import os
import xlrd
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl.reader.excel import load_workbook
from utils import RESOURCES_PATH, TMP_PATH, ZIP_PATH, PDF_FILE, TXT_FILE, XLS_FILE, XLSX_FILE

def test_archive_exists():
    assert os.path.exists(ZIP_PATH)

def test_check_files_in_archive():
    archive_zip_list = ZipFile(ZIP_PATH).namelist()
    assert PDF_FILE in archive_zip_list
    assert TXT_FILE in archive_zip_list
    assert XLS_FILE in archive_zip_list
    assert XLSX_FILE in archive_zip_list
    
def test_file_txt():
    expected_txt_size = os.path.getsize(os.path.join(RESOURCES_PATH, TXT_FILE))
    
    with ZipFile(ZIP_PATH) as archive_zip:
        assert expected_txt_size == archive_zip.getinfo(TXT_FILE).file_size
    
    with open((os.path.join(RESOURCES_PATH, TXT_FILE)), 'r') as file:
        expected_text = file.read().encode('utf-8')           
        
    with ZipFile(ZIP_PATH) as archive_zip:
        actual_text = archive_zip.read(TXT_FILE)
        assert expected_text in actual_text
        
        
def test_pdf():    
    expected_pdf = PdfReader(os.path.join(RESOURCES_PATH, PDF_FILE))

    with ZipFile(ZIP_PATH) as archive_zip:
        with archive_zip.open(PDF_FILE) as zip_pdf_file:
            zip_pdf = PdfReader(zip_pdf_file)
            assert len(expected_pdf.pages) == len(zip_pdf.pages)
            assert expected_pdf.pages[0].extract_text() == zip_pdf.pages[0].extract_text()
        
def test_xls():
    with ZipFile(ZIP_PATH) as archive_zip:
        with archive_zip.open(XLS_FILE) as xls_zip:
            temp = xls_zip.read()
            expected_zip_xls = xlrd.open_workbook(file_contents = temp)
            
        res_xls = xlrd.open_workbook(os.path.join(RESOURCES_PATH, XLS_FILE))
        
        assert expected_zip_xls.nsheets == res_xls.nsheets
        assert expected_zip_xls.sheet_names() == res_xls.sheet_names()

        res_sheet = res_xls.sheet_by_index(0)
        zip_sheet = expected_zip_xls.sheet_by_index(0)
        
        assert res_sheet.ncols == zip_sheet.ncols
        assert res_sheet.nrows == zip_sheet.nrows


def test_xlsx():
    with ZipFile(ZIP_PATH) as hw_zip:
        with hw_zip.open(XLSX_FILE) as xlsx_zip:
            workbook = load_workbook(xlsx_zip)
            sheet = workbook.active
            column = sheet.cell(row=1, column=1).value
            assert column == 'Hello xlsx'
